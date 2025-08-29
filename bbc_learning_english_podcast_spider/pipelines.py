# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

# Standard library imports
from pathlib import Path
import re

# Third-party imports
from scrapy.pipelines.files import FilesPipeline
from openpyxl import Workbook, load_workbook
import scrapy


# ---------------------- Helpers ----------------------
def _sanitize(name: str) -> str:
    """Remove invalid filesystem characters."""

    return re.sub(r'[\\/*?:"<>|]', "", (name or "").strip()) or "Unknown"


# ---------------------- Files Pipeline ----------------------
class CustomFilesPipeline(FilesPipeline):
    """
    Pure-Scrapy media handling:
        - Uses FILES_STORE as root
        - Saves to: <SpiderNameCapitalized>/<Year>/<SafeTitle>.<ext>
        - Skips files if local file exists and sizes match
        - Updates item['pdf_path'] / item['mp3_path']
        - Increments stats for downloaded/skipped
    """

    def get_media_requests(self, item, info):
        store_root = Path(info.spider.crawler.settings.get("FILES_STORE"))
        spider_dir = info.spider.name.capitalize()
        year = item.get("release_year") or "Unknown"

        for ftype in ["pdf_url", "mp3_url"]:
            url = item.get(ftype)
            if url:
                ext = Path(url).suffix.lower()
                title = _sanitize(item.get("title"))
                rel_path = Path(spider_dir) / year / f"{title}{ext}"
                abs_path = store_root / rel_path

                if abs_path.exists() and abs_path.stat().st_size > 0:
                    # skip download
                    info.spider.crawler.stats.inc_value(
                        f"files/{'pdf' if ext=='.pdf' else 'mp3'}_skipped"
                    )
                    # also set the path in the item
                    if ext == ".pdf":
                        item["pdf_path"] = str(abs_path)
                    else:
                        item["mp3_path"] = str(abs_path)
                    continue

                # only yield request if it needs download
                yield scrapy.Request(url, meta={"item": item, "ftype": ftype})

    def file_path(self, request, response=None, info=None, **kwargs):

        # Build a relative path under FILES_STORE
        item = request.meta["item"]
        spider_dir = info.spider.name.capitalize()
        year = item.get("release_year") or "Unknown"
        title = _sanitize(item.get("title"))
        ext = Path(request.url).suffix.lower()

        return f"{spider_dir}/{year}/{title}{ext}"

    def item_completed(self, results, item, info):
        """
        results: list of tuples (success, info_dict)
        info_dict['path'] is the relative path under FILES_STORE
        """
        store = Path(info.spider.crawler.settings.get("FILES_STORE"))

        for ok, finfo in results:
            if not ok:
                continue
            rel_path = Path(finfo["path"])
            abs_path = store / rel_path
            ext = rel_path.suffix.lower()
            if ext == ".pdf":
                item["pdf_path"] = str(abs_path)
                info.spider.crawler.stats.inc_value("files/pdf_downloaded")
            elif ext == ".mp3":
                item["mp3_path"] = str(abs_path)
                info.spider.crawler.stats.inc_value("files/mp3_downloaded")
        return item


class ExcelPipeline:
    """
    Writes rows to Excel located at:
        ~/Documents/BBC_English_Podcast/<SpiderNameCapitalized>/<SpiderNameCapitalized>.xlsx
    Also logs a nice summary with custom logger.
    """

    def open_spider(self, spider):
        store_root = Path(spider.crawler.settings.get("FILES_STORE"))
        self.spider_dir = store_root / spider.name.capitalize()
        self.spider_dir.mkdir(parents=True, exist_ok=True)

        self.excel_path = self.spider_dir / f"{spider.name.capitalize()}.xlsx"
        if self.excel_path.exists():
            self.wb = load_workbook(self.excel_path)
            self.sheet = self.wb.active
        else:
            self.wb = Workbook()
            self.sheet = self.wb.active
            self.sheet.append(
                [
                    "Title",
                    "PDF URL",
                    "PDF Path",
                    "MP3 URL",
                    "MP3 Path",
                    "Page URL",
                    "Release Date",
                    "Release Year",
                    "Status",
                ]
            )
            self.wb.save(self.excel_path)

        self.total_items = 0

    def process_item(self, item, spider):
        self.sheet.append(
            [
                item.get("title"),
                item.get("pdf_url"),
                item.get("pdf_path"),
                item.get("mp3_url"),
                item.get("mp3_path"),
                item.get("url"),
                item.get("release_date"),
                item.get("release_year"),
                "Downloaded",
            ]
        )
        self.total_items += 1

        return item

    def close_spider(self, spider):
        self.wb.save(self.excel_path)

        # Collect stats
        stats = spider.crawler.stats.get_stats()
        total_requests = stats.get("downloader/request_count", 0)
        episodes_scheduled = stats.get("episodes/scheduled", 0)
        episodes_fetched = stats.get("episodes/fetched", 0)
        total_items = stats.get("item_scraped_count", 0)
        pdf_downloaded = stats.get("files/pdf_downloaded", 0)
        mp3_downloaded = stats.get("files/mp3_downloaded", 0)
        pdf_skipped = stats.get("files/pdf_skipped", 0)
        mp3_skipped = stats.get("files/mp3_skipped", 0)

        # Friendly summary with your custom logger
        spider.six_minutes_logger.info(f"\n📁 Excel file saved at {self.excel_path}")
        spider.six_minutes_logger.info(f"==== Crawl Summary ===={'='*20}")
        spider.six_minutes_logger.info(f"🌐 Total requests sent: {total_requests}")
        spider.six_minutes_logger.info(f"📌 Episodes scheduled: {episodes_scheduled}")
        spider.six_minutes_logger.info(
            f"✅ Episodes fetched (HTTP 200): {episodes_fetched}"
        )
        spider.six_minutes_logger.info(f"📝 Total Excel rows: {total_items}")
        spider.six_minutes_logger.info(f"📄 PDFs downloaded: {pdf_downloaded}")
        spider.six_minutes_logger.info(f"⏭️ PDFs skipped (already exist): {pdf_skipped}")
        spider.six_minutes_logger.info(f"🎵 MP3s downloaded: {mp3_downloaded}")
        spider.six_minutes_logger.info(f"⏭️ MP3s skipped (already exist): {mp3_skipped}")
        spider.six_minutes_logger.info(f"========================{'='*20}\n")
